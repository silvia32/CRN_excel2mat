
###############################################################################
# Species extraction from an Excel reaction list
#
# This script reads the list of chemical reactions from the worksheet
# 'table_S3' of an Excel workbook, extracts all unique reactant and product
# species, and creates or updates the worksheet 'table_S1'.
#
# The generated worksheet contains:
#   - Species notation
#   - Species names
#   - Initial concentrations, initialized to 0
#   - Constant-species flags, initialized to 0 and manually customizable
#
# This script must be executed before the Excel-to-SBML conversion script,
# because the generated 'table_S1' worksheet is required to build the
# corresponding SBML model.
#
# Reaction types supported by the complete Excel-to-SBML workflow:
#
# Irreversible reactions:
#   - A -> B
#   - A + B -> C
#   - A -> B + C
#   - A + B -> C + D
#   - A -> B + C + D
#   - A + B -> C + D + E
#
# Reversible reactions:
#   - A <-> B
#   - A + B <-> C
#   - A <-> B + C
#   - A + B <-> C + D
#
# Notes:
# - The complete workflow supports up to two reactants.
# - Irreversible reactions support up to three products.
# - Reversible reactions support up to two products.
# - Duplicate species are automatically removed.
# - Species named 'null' are excluded from the final species list.
# - Initial concentrations and constant-species flags can be customized after
#   generating the 'table_S1' worksheet.
###############################################################################

#%% Import packages
import pandas as pd
import os.path as op
import re
import openpyxl


#%% Load information from Excel file
target = op.join('.', 'excel_data')
file_path = op.join(target, 'smallCRN.xlsx')

reactions_info = pd.read_excel(file_path, sheet_name='table_S3')
n_reactions_forward = len(reactions_info['kf'])


#%% Functions

def extract_first_reactant(string):
    match = re.match(r'(\w+)(?:\s*\+\s*(\w+))?(?:\s*(?:<->|->)\s*(\w+))?', string)
    if match:
        before_plus, _, before_arrow = match.groups()
        return before_plus if before_plus and before_arrow else None
    else:
        return None
    

def extract_second_reactant(string):
    match = re.match(r'\w+\s*\+\s*(\w+)(?:\s*(?:<->|->)\s*(\w+))?', string)
    if match:
        after_plus, _ = match.groups()
        return after_plus
    else:
        return None
    
def extract_first_product(string):
    match = re.search(r'(?:<->|->)\s*(.*?)(?:\s*\+|$)', string, re.IGNORECASE)
    if match:
        return match.group(1)
    else:
        return None

def extract_second_product(string):
    match = re.search(r'(?:<->|->)\s*(\w+)(?:\s*\+\s*(\w+))?', string)
    if match:
        _, after_plus = match.groups()
        return after_plus if after_plus is not None else None
    else:
        return None
    
def extract_third_product(string):
    match = re.search(
        r'(?:<->|->)\s*\w+\s*\+\s*\w+\s*\+\s*(\w+)',
        string
    )
    if match:
        return match.group(1)
    else:
        return None


#%% Extract species
species = []
#products = []

for idx_reaction_row in range(n_reactions_forward):
    reactant1 = extract_first_reactant(reactions_info['Reaction'][idx_reaction_row])
    reactant2 = extract_second_reactant(reactions_info['Reaction'][idx_reaction_row])
    product1 = extract_first_product(reactions_info['Reaction'][idx_reaction_row])
    product2 = extract_second_product(reactions_info['Reaction'][idx_reaction_row])
    product3 = extract_third_product(reactions_info['Reaction'][idx_reaction_row])

    if reactant1 is not None and reactant1 not in species:
        species.append(reactant1)

    if reactant2 is not None and reactant2 not in species:
        species.append(reactant2)

    if product1 is not None and product1 not in species:
        species.append(product1)

    if product2 is not None and product2 not in species:
        species.append(product2)
        
    if product3 is not None and product3 not in species:
        species.append(product3)
        
# Remove 'null' from the species list        
species = [s for s in species if s != 'null']


#%%


n_species = len(species)
species_ids = ['A{}'.format(idx_reaction+1) for idx_reaction in range(n_species)]
initial_concentrations = ['0'] * n_species
is_constant = [0] * n_species

# Only for the mTOR network - customize as needed
for species_name in ['I', 'PIP2gen', 'A']:
    if species_name in species:
        is_constant[species.index(species_name)] = 1

# Create DataFrame from lists
df = pd.DataFrame({'Notation': list(species_ids), 'Protein': list(species), 'x0 [Nm]': list(initial_concentrations), 'IsConstant' : list(is_constant)})

# Open Excel workbook
workbook = openpyxl.load_workbook(file_path)
header = ['Notation', 'Protein', 'x0 [Nm]', 'IsConstant']

if 'table_S1' in workbook.sheetnames:
    print(workbook.sheetnames)
    existing_sheet = workbook['table_S1']
    for row in existing_sheet.iter_rows(min_row=1, max_row=existing_sheet.max_row, min_col=1, max_col=existing_sheet.max_column):
        for cell in row:
            cell.value = None
    for c_idx, value in enumerate(header, start=1):
        existing_sheet.cell(row=1, column=c_idx, value=value)
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            existing_sheet.cell(row=r_idx, column=c_idx, value=value)
else:
    print('Worksheet does not exist')
    new_sheet = workbook.create_sheet('table_S1')
    for c_idx, value in enumerate(header, start=1):
        new_sheet.cell(row=1, column=c_idx, value=value)
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            new_sheet.cell(row=r_idx, column=c_idx, value=value)

workbook.save(file_path)
workbook.close()


# %%
